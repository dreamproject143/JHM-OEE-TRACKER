# app.py
import os
import pandas as pd
import numpy as np
from flask import Flask, render_template, request, send_file, jsonify, send_from_directory
from openpyxl import load_workbook
from io import BytesIO

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed'

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'files' not in request.files:
        return jsonify({'message': 'No files uploaded', 'status': 'error'})
    
    files = request.files.getlist('files')
    if len(files) == 0:
        return jsonify({'message': 'No selected files', 'status': 'error'})

    # Clear existing files in upload folder
    for existing_file in os.listdir(UPLOAD_FOLDER):
        file_path = os.path.join(UPLOAD_FOLDER, existing_file)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
        except Exception as e:
            print(f"Error deleting {file_path}: {e}")

    # Save new files
    for file in files:
        if file.filename == '':
            continue
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(file_path)

    return jsonify({
        'message': f'{len(files)} files uploaded successfully!',
        'status': 'success',
        'count': len(files)
    })

def process_files():
    machine_data = {}
    files = os.listdir(UPLOAD_FOLDER)
    
    for file_name in files:
        file_path = os.path.join(UPLOAD_FOLDER, file_name)
        print(f"📂 Processing file: {file_name}")

        try:
            wb = load_workbook(filename=file_path, data_only=True)
            if "MANUAL LINE" not in wb.sheetnames:
                print(f"⚠️ Skipping {file_name}: No 'MANUAL LINE' sheet found.")
                continue

            sheet = wb['MANUAL LINE']
            header_row = None
            gross_col = None

            for row in sheet.iter_rows():
                cell_value = row[0].value
                if cell_value and str(cell_value).strip().lower() == "work center":
                    header_row = row[0].row
                    for idx, cell in enumerate(sheet[header_row]):
                        if cell.value and "gross" in str(cell.value).lower():
                            gross_col = idx
                            break
                    break

            if header_row is None or gross_col is None:
                print(f"⚠️ Skipping {file_name}: Header or Gross column not found.")
                continue

            date_row = header_row - 1
            schedule_date = "Unknown"
            if date_row >= 1:
                date_cell = sheet.cell(row=date_row, column=1)
                raw_date = date_cell.value
                schedule_date = pd.to_datetime(raw_date, errors='coerce').strftime('%d-%b') if raw_date else "Unknown"

            for row in sheet.iter_rows(min_row=header_row + 1):
                machine_cell = row[0]
                gross_cell = row[gross_col]

                machine = str(machine_cell.value).strip() if machine_cell.value else ""
                if not machine:
                    continue

                gross_value = gross_cell.value
                if gross_value is None:
                    continue

                percentage = None
                if isinstance(gross_value, str):
                    clean_value = gross_value.replace('%', '').strip()
                    try: percentage = float(clean_value) / 100
                    except: continue
                else:
                    percentage = gross_value / 100 if not gross_cell.number_format or '%' not in gross_cell.number_format else gross_value

                try: percentage = float(percentage)
                except: continue

                if machine not in machine_data:
                    machine_data[machine] = {}
                if schedule_date not in machine_data[machine]:
                    machine_data[machine][schedule_date] = []
                machine_data[machine][schedule_date].append(percentage)

        except Exception as e:
            print(f"❌ Error reading {file_name}: {str(e)}")

    # Calculate averages
    machine_performance = {}
    for machine, dates in machine_data.items():
        machine_performance[machine] = {}
        for date, values in dates.items():
            if values: machine_performance[machine][date] = sum(values) / len(values)
        all_values = [v for date_values in dates.values() for v in date_values]
        if all_values: machine_performance[machine]['Average'] = sum(all_values) / len(all_values)

    # Create DataFrame and ensure 'Average' is last column
    summary_df = pd.DataFrame(machine_performance).T.sort_index(ascending=True)
    
    # Reorder columns to place 'Average' last
    cols = summary_df.columns.tolist()
    if 'Average' in cols:
        cols.remove('Average')
        cols.append('Average')
        summary_df = summary_df[cols]
    
    summary_df.index.name = "Work Center"
    return summary_df

@app.route('/process', methods=['POST'])
def process_and_download():
    try:
        summary_df = process_files()
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            summary_df.to_excel(writer, sheet_name="Summary", na_rep='')
            workbook = writer.book
            worksheet = writer.sheets['Summary']

            # Formatting
            header_format = workbook.add_format({
                'bold': True, 'font_size': 12, 'bg_color': '#1F497D',
                'font_color': 'white', 'border': 1, 'align': 'center'
            })
            
            cell_format = workbook.add_format({
                'border': 1, 'align': 'center', 'num_format': '0.00%',
                'font_color': '#1F497D', 'bg_color': '#FFFFFF'
            })

            # Write header
            worksheet.write_row(0, 0, ["Work Center"] + summary_df.columns.tolist(), header_format)

            # Write data
            for row_idx, (index, row) in enumerate(summary_df.iterrows(), start=1):
                worksheet.write(row_idx, 0, index, cell_format)
                for col_idx, value in enumerate(row, start=1):
                    if pd.notna(value):
                        worksheet.write(row_idx, col_idx, value, cell_format)

            # Adjust column widths
            worksheet.set_column(0, 0, 20)  # Work Center column
            for idx, col in enumerate(summary_df.columns, start=1):
                worksheet.set_column(idx, idx, 15)

        output.seek(0)
        return send_file(output, download_name='Machine_Performance_Summary.xlsx', as_attachment=True)

    except Exception as e:
        return jsonify({'message': f'Error processing files: {str(e)}', 'status': 'error'})

if __name__ == '__main__':
    app.run(debug=True)
